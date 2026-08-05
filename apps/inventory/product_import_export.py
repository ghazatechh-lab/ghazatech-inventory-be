from __future__ import annotations

from collections import OrderedDict
from io import BytesIO
from typing import Any

from django.db import transaction
from django.http import FileResponse
from openpyxl import Workbook, load_workbook
from rest_framework import serializers

from .models import Brand, Category, Product, Rack
from .serializers import ProductSerializer

EXPECTED_HEADERS = [
    "product name",
    "sku",
    "barcode",
    "brand",
    "category",
    "supplier",
    "rack",
    "compatible model",
    "condition",
    "warranty period",
    "reorder level",
    "attribute",
    "value",
    "initial quantity",
    "purchase price",
    "retail price",
    "wholesale price",
    "minimum selling",
]


def _normalise_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _normalise_text(value: Any) -> str:
    return str(value or "").strip()


def _as_integer(value: Any, *, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise serializers.ValidationError(
            f"Expected a whole number, received {value!r}."
        ) from exc


def _as_decimal_or_none(value: Any):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise serializers.ValidationError(
            f"Expected a numeric value, received {value!r}."
        ) from exc


def _warranty_days(value: Any) -> int:
    if value in (None, ""):
        return 0

    text = _normalise_text(value).upper()
    number = ""

    for character in text:
        if character.isdigit() or character == ".":
            number += character
        elif number:
            break

    if not number:
        raise serializers.ValidationError(
            f"Warranty period {value!r} must include a number of days, months or years."
        )

    amount = float(number)

    if "YEAR" in text:
        return int(round(amount * 365))
    if "MONTH" in text:
        return int(round(amount * 30))

    return int(round(amount))


def _get_by_name(model, field_name: str, value: Any, row_number: int):
    name = _normalise_text(value)

    if not name:
        raise serializers.ValidationError(
            {f"row_{row_number}": [f"{model._meta.verbose_name.title()} is required."]}
        )

    queryset = model.objects.filter(**{f"{field_name}__iexact": name})

    if model in {Brand, Category}:
        queryset = queryset.filter(is_active=True)

    item = queryset.first()

    if not item:
        raise serializers.ValidationError(
            {
                f"row_{row_number}": [
                    f"{model._meta.verbose_name.title()} '{name}' was not found."
                ]
            }
        )

    return item


def _get_supplier(value: Any, row_number: int):
    name = _normalise_text(value)
    if not name:
        return None

    from apps.suppliers.models import Supplier

    item = Supplier.objects.filter(supplier_name__iexact=name).first()
    if not item:
        raise serializers.ValidationError(
            {f"row_{row_number}": [f"Supplier '{name}' was not found."]}
        )
    return item


def _get_rack(branch, value: Any, row_number: int):
    code = _normalise_text(value)
    if not code:
        return None

    item = Rack.objects.filter(
        branch=branch,
        rack_code__iexact=code,
        is_active=True,
    ).first()

    if not item:
        raise serializers.ValidationError(
            {
                f"row_{row_number}": [
                    f"Rack '{code}' was not found in branch {branch.branch_code}."
                ]
            }
        )

    return item


def parse_product_workbook(uploaded_file) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise serializers.ValidationError(
            {"file": ["Unable to read the Excel workbook. Upload a valid .xlsx file."]}
        ) from exc

    sheet = workbook.active
    raw_headers = [_normalise_header(cell.value) for cell in sheet[1]]
    header_indexes = {header: index for index, header in enumerate(raw_headers)}

    missing = [header for header in EXPECTED_HEADERS if header not in header_indexes]
    if missing:
        raise serializers.ValidationError(
            {"file": ["The template is missing these columns: " + ", ".join(missing)]}
        )

    grouped: list[dict[str, Any]] = []
    current_product = None

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):

        def value(header):
            return row[header_indexes[header]]

        product_name = _normalise_text(value("product name"))
        sku = _normalise_text(value("sku"))

        row_has_data = any(cell not in (None, "") for cell in row)
        if not row_has_data:
            continue

        if product_name or sku:
            if not product_name or not sku:
                raise serializers.ValidationError(
                    {
                        f"row_{row_number}": [
                            "Both product name and SKU are required on a product row."
                        ]
                    }
                )

            current_product = {
                "row_number": row_number,
                "product_name": product_name,
                "sku": sku,
                "barcode": value("barcode"),
                "brand": value("brand"),
                "category": value("category"),
                "supplier": value("supplier"),
                "rack": value("rack"),
                "compatible_models": _normalise_text(value("compatible model"))
                or "All Models",
                "condition": _normalise_text(value("condition")) or "NEW",
                "warranty_period_days": _warranty_days(value("warranty period")),
                "reorder_level": _as_integer(value("reorder level"), default=0),
                "variants": [],
            }
            grouped.append(current_product)
        elif current_product is None:
            raise serializers.ValidationError(
                {f"row_{row_number}": ["A continuation row must follow a product row."]}
            )

        attribute_name = _normalise_text(value("attribute"))
        attribute_value = _normalise_text(value("value"))
        initial_quantity = _as_integer(value("initial quantity"), default=0)
        purchase_price = _as_decimal_or_none(value("purchase price"))
        retail_price = _as_decimal_or_none(value("retail price")) or 0
        wholesale_price = _as_decimal_or_none(value("wholesale price")) or 0
        minimum_selling_price = _as_decimal_or_none(value("minimum selling")) or 0

        has_variant_data = any(
            item not in ("", None, 0)
            for item in (
                attribute_name,
                attribute_value,
                initial_quantity,
                purchase_price,
                retail_price,
                wholesale_price,
                minimum_selling_price,
            )
        )

        if not has_variant_data:
            continue

        if bool(attribute_name) != bool(attribute_value):
            raise serializers.ValidationError(
                {f"row_{row_number}": ["Attribute and value must both be provided."]}
            )

        current_product["variants"].append(
            {
                "attributes": (
                    {attribute_name: attribute_value}
                    if attribute_name and attribute_value
                    else {}
                ),
                "initial_stock": max(0, initial_quantity),
                "purchase_price": purchase_price,
                "retail_price": retail_price,
                "wholesale_price": wholesale_price,
                "minimum_selling_price": minimum_selling_price,
                "is_active": True,
            }
        )

    if not grouped:
        raise serializers.ValidationError(
            {"file": ["The workbook does not contain any product rows."]}
        )

    return grouped


@transaction.atomic
def import_products_from_workbook(*, request, branch, uploaded_file):
    grouped = parse_product_workbook(uploaded_file)
    created_products = []

    for product_data in grouped:
        row_number = product_data.pop("row_number")
        brand_name = product_data.pop("brand")
        category_name = product_data.pop("category")
        supplier_name = product_data.pop("supplier")
        rack_code = product_data.pop("rack")
        variants = product_data.pop("variants")

        brand = _get_by_name(Brand, "name", brand_name, row_number)
        category = _get_by_name(Category, "name", category_name, row_number)
        supplier = _get_supplier(supplier_name, row_number)
        rack = _get_rack(branch, rack_code, row_number)

        sku = product_data["sku"]

        if Product.objects.filter(
            branch=branch,
            sku__iexact=sku,
            is_deleted=False,
        ).exists():
            raise serializers.ValidationError(
                {
                    f"row_{row_number}": [
                        f"SKU '{sku}' already exists in branch {branch.branch_code}."
                    ]
                }
            )

        barcode = product_data.get("barcode")
        if barcode not in (None, ""):
            barcode = str(barcode).strip()
            if Product.objects.filter(barcode=barcode).exists():
                raise serializers.ValidationError(
                    {f"row_{row_number}": [f"Barcode '{barcode}' already exists."]}
                )
        else:
            barcode = None

        if not variants:
            variants = [
                {
                    "attributes": {},
                    "initial_stock": 0,
                    "purchase_price": None,
                    "retail_price": 0,
                    "wholesale_price": 0,
                    "minimum_selling_price": 0,
                    "is_active": True,
                }
            ]

        has_variants = any(bool(item.get("attributes")) for item in variants)

        payload = {
            **product_data,
            "barcode": barcode,
            "brand": brand.id,
            "category": category.id,
            "supplier": supplier.id if supplier else None,
            "branch": branch.id,
            "rack": rack.id if rack else None,
            "has_variants": has_variants,
            "unit": "PCS",
            "tax_treatment": "VAT",
            "vat_inclusive": True,
            "description": "",
            "is_active": True,
            "variants": variants,
        }

        serializer = ProductSerializer(
            data=payload,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        created_products.append(serializer.save())

    return created_products


def build_product_export_workbook(queryset) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"

    headers = [
        "product name",
        "sku",
        "barcode",
        "brand",
        "category",
        "supplier",
        "rack",
        "compatible model",
        "condition",
        "warranty period",
        "reorder level",
        "attribute ",
        "value ",
        "Initial quantity",
        "purchase price",
        "retail price",
        "wholesale price",
        "minimum selling",
    ]
    sheet.append(headers)

    for product in queryset:
        variants = list(product.variants.filter(is_active=True).order_by("id"))
        if not variants:
            variants = [None]

        for variant_index, variant in enumerate(variants):
            attribute_items = (
                list((variant.attributes or {}).items()) if variant else []
            )
            if not attribute_items:
                attribute_items = [("", "")]

            for attribute_index, (attribute_name, attribute_value) in enumerate(
                attribute_items
            ):
                first_row = variant_index == 0 and attribute_index == 0
                branch_id = product.branch_id
                stock_variant = variant if product.has_variants else None
                stock = product.stocks.filter(
                    branch_id=branch_id,
                    variant=stock_variant,
                ).first()

                sheet.append(
                    [
                        product.product_name if first_row else "",
                        product.sku if first_row else "",
                        product.barcode if first_row else "",
                        product.brand.name if first_row else "",
                        product.category.name if first_row else "",
                        (
                            product.supplier.supplier_name
                            if first_row and product.supplier_id
                            else ""
                        ),
                        (
                            product.rack.rack_code
                            if first_row and product.rack_id
                            else ""
                        ),
                        (
                            (product.compatible_models or "All Models")
                            if first_row
                            else ""
                        ),
                        product.condition if first_row else "",
                        (
                            f"{product.warranty_period_days} DAYS"
                            if first_row and product.warranty_period_days
                            else ""
                        ),
                        product.reorder_level if first_row else "",
                        attribute_name,
                        attribute_value,
                        int(stock.current_stock or 0) if stock else 0,
                        variant.purchase_price if variant else "",
                        variant.retail_price if variant else 0,
                        variant.wholesale_price if variant else 0,
                        variant.minimum_selling_price if variant else 0,
                    ]
                )

    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)
        cell.fill = cell.fill.copy(fill_type="solid", fgColor="D9EAF7")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 30,
        "B": 18,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 20,
        "G": 18,
        "H": 24,
        "I": 14,
        "J": 18,
        "K": 14,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def product_export_response(queryset):
    output = build_product_export_workbook(queryset)
    return FileResponse(
        output,
        as_attachment=True,
        filename="products_export.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
