import csv
from io import BytesIO
from datetime import timedelta
from decimal import Decimal

from django.db import transaction
from django.db.models import Avg, Count, DurationField, ExpressionWrapper, F, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.common.logging import LoggedModelViewSet as ModelViewSet

from .models import *
from apps.inventory.models import Product, ProductStock, StockMovement
from apps.finance.models import BankAccount, CashRegister
from apps.customers.models import Customer
from apps.branches.models import Branch
from apps.accounts.models import User
from .serializers import *


class Base(ModelViewSet):
    def get_queryset(self):
        qs = super().get_queryset()
        b = self.request.query_params.get("branch")
        return qs.filter(branch_id=b) if b else qs


class QuotationViewSet(Base):
    queryset = Quotation.objects.select_related(
        "customer",
        "branch",
        "salesperson",
    ).prefetch_related(
        "items__product",
        "items__variant",
    )

    serializer_class = QuotationSerializer

    search_fields = [
        "quote_number",
        "customer__customer_name",
        "customer__phone",
        "customer__email",
        "status",
        "payment_terms",
        "delivery_terms",
    ]

    filterset_fields = [
        "branch",
        "customer",
        "salesperson",
        "status",
        "currency",
    ]

    ordering_fields = [
        "quote_number",
        "quote_date",
        "valid_until",
        "total_amount",
        "status",
        "created_at",
        "customer__customer_name",
    ]

    ordering = [
        "-quote_date",
        "-id",
    ]

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        today = timezone.localdate()
        week_start = today - timedelta(days=today.weekday())

        open_statuses = [
            "DRAFT",
            "SENT",
            "PENDING",
        ]

        accepted_this_month = queryset.filter(
            status="ACCEPTED",
            accepted_at__year=today.year,
            accepted_at__month=today.month,
        )

        turnaround = (
            accepted_this_month.exclude(accepted_at=None)
            .annotate(
                turnaround=ExpressionWrapper(
                    F("accepted_at") - F("created_at"),
                    output_field=DurationField(),
                )
            )
            .aggregate(value=Avg("turnaround"))["value"]
        )

        avg_days = (
            round(
                turnaround.total_seconds() / 86400,
                1,
            )
            if turnaround
            else 0
        )

        created_this_week = queryset.filter(
            created_at__date__gte=week_start,
        ).count()

        return Response(
            {
                "open_quotations": queryset.filter(
                    status__in=open_statuses,
                ).count(),
                "open_change": f"+{created_this_week}",
                "value_pending": queryset.filter(
                    status__in=[
                        "SENT",
                        "PENDING",
                    ]
                ).aggregate(value=Sum("total_amount"))["value"]
                or 0,
                "accepted_this_month": accepted_this_month.count(),
                "acceptance_change": 0,
                "avg_turnaround_days": avg_days,
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = 'attachment; filename="quotations.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Quote Number",
                "Customer",
                "Branch",
                "Quote Date",
                "Valid Until",
                "Currency",
                "Subtotal",
                "VAT",
                "Discount",
                "Total",
                "Status",
            ]
        )

        for quotation in queryset:
            writer.writerow(
                [
                    quotation.quote_number,
                    (quotation.customer.customer_name if quotation.customer else ""),
                    (quotation.branch.branch_name if quotation.branch else ""),
                    quotation.quote_date or "",
                    quotation.valid_until or "",
                    quotation.currency or "AED",
                    quotation.subtotal or 0,
                    quotation.vat_amount or 0,
                    quotation.discount_amount or 0,
                    quotation.total_amount or 0,
                    quotation.get_status_display(),
                ]
            )

        return response

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
        url_path="convert-to-order",
    )
    def convert_to_order(self, request, pk=None):
        quotation = self.get_object()

        if quotation.status == "CONVERTED":
            existing_order = quotation.sales_orders.order_by("-id").first()

            if existing_order:
                return Response(
                    SalesOrderSerializer(
                        existing_order,
                        context={
                            "request": request,
                        },
                    ).data
                )

        if quotation.status in [
            "REJECTED",
            "EXPIRED",
        ]:
            return Response(
                {"detail": "Rejected or expired quotations cannot be converted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order_prefix = timezone.now().strftime("SO-%Y%m")

        order_count = SalesOrder.objects.filter(
            order_number__startswith=order_prefix,
        ).count()

        order = SalesOrder.objects.create(
            order_number=f"{order_prefix}-{order_count + 1:04d}",
            quotation=quotation,
            customer=quotation.customer,
            branch=quotation.branch,
            salesperson=quotation.salesperson,
            order_date=timezone.localdate(),
            currency=quotation.currency,
            subtotal=quotation.subtotal,
            discount_amount=quotation.discount_amount,
            vat_amount=quotation.vat_amount,
            shipping_amount=quotation.shipping_amount,
            total_amount=quotation.total_amount,
            notes=quotation.notes,
            status="DRAFT",
        )

        for item in quotation.items.all():
            order.items.create(
                product=item.product,
                variant=item.variant,
                description=item.description,
                quantity=item.quantity,
                unit_price=item.unit_price,
                vat_percentage=item.vat_percentage,
                line_total=item.line_total,
            )

        quotation.status = "CONVERTED"
        quotation.converted_at = timezone.now()

        quotation.save(
            update_fields=[
                "status",
                "converted_at",
                "updated_at",
            ]
        )

        return Response(
            SalesOrderSerializer(
                order,
                context={
                    "request": request,
                },
            ).data,
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=["post"],
    )
    def send(self, request, pk=None):
        quotation = self.get_object()

        if quotation.status in [
            "CONVERTED",
            "REJECTED",
            "EXPIRED",
        ]:
            return Response(
                {"detail": "This quotation cannot be sent."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        quotation.status = "SENT"
        quotation.sent_at = timezone.now()

        quotation.save(
            update_fields=[
                "status",
                "sent_at",
                "updated_at",
            ]
        )

        return Response(
            self.get_serializer(
                quotation,
            ).data
        )

    @action(
        detail=True,
        methods=["post"],
    )
    def accept(self, request, pk=None):
        quotation = self.get_object()

        quotation.status = "ACCEPTED"
        quotation.accepted_at = timezone.now()

        quotation.save(
            update_fields=[
                "status",
                "accepted_at",
                "updated_at",
            ]
        )

        return Response(
            self.get_serializer(
                quotation,
            ).data
        )

    @action(
        detail=True,
        methods=["post"],
    )
    def reject(self, request, pk=None):
        quotation = self.get_object()

        quotation.status = "REJECTED"
        quotation.rejected_at = timezone.now()

        quotation.save(
            update_fields=[
                "status",
                "rejected_at",
                "updated_at",
            ]
        )

        return Response(
            self.get_serializer(
                quotation,
            ).data
        )


class SalesOrderViewSet(Base):
    queryset = SalesOrder.objects.select_related(
        "customer",
        "branch",
        "salesperson",
        "quotation",
    ).prefetch_related(
        "items__product",
        "items__variant",
    )

    serializer_class = SalesOrderSerializer

    search_fields = [
        "order_number",
        "customer__customer_name",
        "quotation__quote_number",
        "status",
        "shipping_address",
        "emirate",
    ]

    filterset_fields = [
        "branch",
        "customer",
        "salesperson",
        "quotation",
        "status",
        "delivery_method",
    ]

    ordering_fields = [
        "order_number",
        "order_date",
        "delivery_date",
        "total_amount",
        "status",
        "created_at",
        "customer__customer_name",
    ]

    ordering = [
        "-order_date",
        "-id",
    ]

    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        branches = Branch.objects.filter(
            is_active=True,
        ).order_by("branch_name")

        customers = Customer.objects.filter(
            is_active=True,
        ).order_by("customer_name")

        salespeople = User.objects.filter(
            is_active=True,
        ).order_by("first_name", "username")

        products = Product.objects.filter(
            is_active=True,
        ).order_by("product_name")

        quotations = (
            Quotation.objects.select_related("customer", "branch")
            .filter(
                status__in=[
                    "ACCEPTED",
                    "SENT",
                    "PENDING",
                ]
            )
            .order_by("-quote_date", "-id")
        )

        stocks = ProductStock.objects.select_related(
            "product",
            "variant",
            "branch",
        )

        if branch_id:
            quotations = quotations.filter(
                branch_id=branch_id,
            )
            stocks = stocks.filter(
                branch_id=branch_id,
            )

        return Response(
            {
                "branches": [
                    {
                        "id": branch.id,
                        "branch_name": branch.branch_name,
                        "branch_code": getattr(branch, "branch_code", ""),
                        "location": getattr(branch, "address", ""),
                    }
                    for branch in branches
                ],
                "customers": [
                    {
                        "id": customer.id,
                        "customer_name": customer.customer_name,
                    }
                    for customer in customers
                ],
                "salespeople": [
                    {
                        "id": user.id,
                        "display_name": (user.get_full_name() or user.username),
                    }
                    for user in salespeople
                ],
                "products": [
                    {
                        "id": product.id,
                        "product_name": product.product_name,
                        "sku": getattr(product, "sku", ""),
                        "description": getattr(product, "description", ""),
                        "selling_price": getattr(product, "selling_price", 0),
                    }
                    for product in products
                ],
                "quotations": [
                    {
                        "id": quotation.id,
                        "quote_number": quotation.quote_number,
                        "customer_name": (
                            quotation.customer.customer_name
                            if quotation.customer
                            else ""
                        ),
                        "total_amount": quotation.total_amount,
                    }
                    for quotation in quotations
                ],
                "stock": [
                    {
                        "product_id": stock.product_id,
                        "variant_id": stock.variant_id,
                        "branch_id": stock.branch_id,
                        "available_stock": (stock.current_stock - stock.reserved_stock),
                    }
                    for stock in stocks
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        today = timezone.localdate()

        open_statuses = [
            "DRAFT",
            "PENDING",
            "CONFIRMED",
            "AWAITING_FULFILLMENT",
            "PARTIALLY_FULFILLED",
        ]

        month_queryset = queryset.filter(
            order_date__year=today.year,
            order_date__month=today.month,
        )

        fulfilled = queryset.filter(
            status="FULFILLED",
        )

        fulfilled_on_time = fulfilled.filter(
            fulfilled_at__date__lte=F("delivery_date"),
        ).count()

        fulfilled_count = fulfilled.count()

        return Response(
            {
                "open_orders": queryset.filter(
                    status__in=open_statuses,
                ).count(),
                "open_today": queryset.filter(
                    created_at__date=today,
                ).count(),
                "awaiting_fulfillment": queryset.filter(
                    status__in=[
                        "CONFIRMED",
                        "AWAITING_FULFILLMENT",
                        "PARTIALLY_FULFILLED",
                    ]
                ).count(),
                "order_value_mtd": month_queryset.aggregate(value=Sum("total_amount"))[
                    "value"
                ]
                or 0,
                "order_value_change": 0,
                "fulfilled_on_time": (
                    round(
                        fulfilled_on_time / fulfilled_count * 100,
                        1,
                    )
                    if fulfilled_count
                    else 0
                ),
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = 'attachment; filename="sales-orders.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Order Number",
                "Customer",
                "Quotation",
                "Branch",
                "Order Date",
                "Delivery Date",
                "Delivery Method",
                "Currency",
                "Total",
                "Status",
            ]
        )

        for order in queryset:
            writer.writerow(
                [
                    order.order_number,
                    (order.customer.customer_name if order.customer else ""),
                    (order.quotation.quote_number if order.quotation else ""),
                    (order.branch.branch_name if order.branch else ""),
                    order.order_date or "",
                    order.delivery_date or "",
                    order.get_delivery_method_display(),
                    order.currency or "AED",
                    order.total_amount or 0,
                    order.get_status_display(),
                ]
            )

        return response

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
    )
    def confirm(self, request, pk=None):
        order = self.get_object()

        if order.status != "DRAFT":
            return Response(
                {"detail": "Only draft sales orders can be confirmed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for item in order.items.all():
            stock = (
                ProductStock.objects.select_for_update()
                .filter(
                    product=item.product,
                    variant=item.variant,
                    branch=order.branch,
                )
                .first()
            )

            if not stock:
                return Response(
                    {
                        "detail": f"No stock record found for {item.product.product_name}."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            available = stock.current_stock - stock.reserved_stock

            if available < item.quantity:
                return Response(
                    {"detail": f"Insufficient stock for {item.product.product_name}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            stock.reserved_stock += item.quantity
            stock.save(
                update_fields=[
                    "reserved_stock",
                    "updated_at",
                ]
            )

        order.status = "CONFIRMED"
        order.confirmed_at = timezone.now()

        order.save(
            update_fields=[
                "status",
                "confirmed_at",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(order).data)

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
        url_path="convert-to-invoice",
    )
    def convert_to_invoice(self, request, pk=None):
        order = self.get_object()

        existing_invoice = (
            order.invoices.exclude(payment_status="VOID").order_by("-id").first()
        )

        if existing_invoice:
            return Response(
                SalesInvoiceSerializer(
                    existing_invoice,
                    context={"request": request},
                ).data
            )

        if order.status == "CANCELLED":
            return Response(
                {"detail": "Cancelled orders cannot be invoiced."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        prefix = timezone.now().strftime("INV-%Y%m")
        count = SalesInvoice.objects.filter(
            invoice_number__startswith=prefix,
        ).count()

        invoice = SalesInvoice.objects.create(
            invoice_number=f"{prefix}-{count + 1:04d}",
            sales_order=order,
            customer=order.customer,
            branch=order.branch,
            salesperson=order.salesperson,
            invoice_date=timezone.localdate(),
            due_date=order.delivery_date,
            currency=order.currency,
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            vat_amount=order.vat_amount,
            shipping_amount=order.shipping_amount,
            total_amount=order.total_amount,
            balance_due=order.total_amount,
            sale_type="ORDER",
            payment_status="UNPAID",
            notes=order.notes,
        )

        for item in order.items.all():
            invoice.items.create(
                sales_order_item=item,
                product=item.product,
                variant=item.variant,
                description=item.description,
                quantity=item.quantity,
                unit_price=item.unit_price,
                vat_percentage=item.vat_percentage,
                line_total=item.line_total,
            )

        return Response(
            SalesInvoiceSerializer(
                invoice,
                context={"request": request},
            ).data,
            status=status.HTTP_201_CREATED,
        )


class SalesInvoiceViewSet(Base):
    queryset = SalesInvoice.objects.select_related(
        "customer",
        "branch",
        "sales_order",
        "salesperson",
        "bank_account",
    ).prefetch_related(
        "items__product",
        "items__variant",
        "items__sales_order_item",
        "payments",
    )

    serializer_class = SalesInvoiceSerializer

    search_fields = [
        "invoice_number",
        "customer__customer_name",
        "sales_order__order_number",
        "customer_po_number",
        "payment_status",
    ]

    filterset_fields = [
        "branch",
        "customer",
        "salesperson",
        "sales_order",
        "payment_status",
        "sale_type",
        "payment_terms",
    ]

    ordering_fields = [
        "invoice_number",
        "invoice_date",
        "due_date",
        "total_amount",
        "paid_amount",
        "balance_due",
        "payment_status",
        "created_at",
        "customer__customer_name",
    ]

    ordering = [
        "-invoice_date",
        "-id",
    ]

    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        branches = Branch.objects.filter(
            is_active=True,
        ).order_by("branch_name")

        customers = Customer.objects.filter(
            is_active=True,
        ).order_by("customer_name")

        salespeople = User.objects.filter(
            is_active=True,
        ).order_by("first_name", "username")

        product_filters = {
            "is_active": True,
        }

        product_field_names = {field.name for field in Product._meta.get_fields()}

        if "is_deleted" in product_field_names:
            product_filters["is_deleted"] = False

        products = Product.objects.filter(
            **product_filters,
        ).order_by("product_name")

        sales_orders = (
            SalesOrder.objects.select_related(
                "customer",
                "branch",
            )
            .exclude(
                status__in=[
                    "CANCELLED",
                ]
            )
            .order_by(
                "-order_date",
                "-id",
            )
        )

        bank_accounts = BankAccount.objects.filter(
            is_active=True,
        ).order_by("account_name")

        if branch_id:
            sales_orders = sales_orders.filter(
                branch_id=branch_id,
            )

            bank_account_fields = {
                field.name for field in BankAccount._meta.get_fields()
            }

            if "branch" in bank_account_fields:
                bank_accounts = bank_accounts.filter(
                    branch_id=branch_id,
                )

        return Response(
            {
                "branches": [
                    {
                        "id": branch.id,
                        "branch_name": branch.branch_name,
                        "branch_code": getattr(
                            branch,
                            "branch_code",
                            "",
                        ),
                        "location": getattr(
                            branch,
                            "address",
                            "",
                        ),
                        "trn": getattr(
                            branch,
                            "trn",
                            "",
                        ),
                    }
                    for branch in branches
                ],
                "customers": [
                    {
                        "id": customer.id,
                        "customer_name": customer.customer_name,
                    }
                    for customer in customers
                ],
                "salespeople": [
                    {
                        "id": user.id,
                        "display_name": (user.get_full_name() or user.username),
                    }
                    for user in salespeople
                ],
                "products": [
                    {
                        "id": product.id,
                        "product_name": product.product_name,
                        "sku": getattr(
                            product,
                            "sku",
                            "",
                        ),
                        "description": getattr(
                            product,
                            "description",
                            "",
                        ),
                        "selling_price": getattr(
                            product,
                            "selling_price",
                            0,
                        ),
                        "is_active": getattr(
                            product,
                            "is_active",
                            True,
                        ),
                        "is_deleted": getattr(
                            product,
                            "is_deleted",
                            False,
                        ),
                    }
                    for product in products
                ],
                "sales_orders": [
                    {
                        "id": order.id,
                        "order_number": order.order_number,
                        "customer_name": (
                            order.customer.customer_name if order.customer else ""
                        ),
                        "total_amount": order.total_amount,
                    }
                    for order in sales_orders
                ],
                "bank_accounts": [
                    {
                        "id": account.id,
                        "account_name": getattr(
                            account,
                            "account_name",
                            str(account),
                        ),
                        "bank_name": getattr(
                            account,
                            "bank_name",
                            "",
                        ),
                        "iban": getattr(
                            account,
                            "iban",
                            "",
                        ),
                    }
                    for account in bank_accounts
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        today = timezone.localdate()

        outstanding_queryset = queryset.filter(
            balance_due__gt=0,
        ).exclude(
            payment_status="VOID",
        )

        overdue_queryset = outstanding_queryset.filter(
            due_date__lt=today,
        )

        paid_this_month_queryset = queryset.filter(
            payment_status="PAID",
            paid_at__year=today.year,
            paid_at__month=today.month,
        )

        paid_invoices = queryset.filter(
            payment_status="PAID",
            paid_at__isnull=False,
            invoice_date__isnull=False,
        )

        durations = []

        for invoice in paid_invoices:
            paid_date = invoice.paid_at.date()

            durations.append((paid_date - invoice.invoice_date).days)

        average_days = (
            round(
                sum(durations) / len(durations),
                1,
            )
            if durations
            else 0
        )

        return Response(
            {
                "outstanding": outstanding_queryset.aggregate(value=Sum("balance_due"))[
                    "value"
                ]
                or 0,
                "outstanding_count": outstanding_queryset.count(),
                "overdue": overdue_queryset.aggregate(value=Sum("balance_due"))["value"]
                or 0,
                "overdue_count": overdue_queryset.count(),
                "paid_this_month": paid_this_month_queryset.aggregate(
                    value=Sum("total_amount")
                )["value"]
                or 0,
                "paid_change": 0,
                "avg_days_to_pay": average_days,
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = 'attachment; filename="sales-invoices.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Invoice Number",
                "Customer",
                "Sales Order",
                "Branch",
                "Issue Date",
                "Due Date",
                "Currency",
                "Total",
                "Paid",
                "Balance Due",
                "Status",
            ]
        )

        for invoice in queryset:
            writer.writerow(
                [
                    invoice.invoice_number,
                    (invoice.customer.customer_name if invoice.customer else ""),
                    (invoice.sales_order.order_number if invoice.sales_order else ""),
                    (invoice.branch.branch_name if invoice.branch else ""),
                    invoice.invoice_date or "",
                    invoice.due_date or "",
                    invoice.currency or "AED",
                    invoice.total_amount or 0,
                    invoice.paid_amount or 0,
                    invoice.balance_due or 0,
                    invoice.get_payment_status_display(),
                ]
            )

        return response

    @action(
        detail=True,
        methods=["post"],
        url_path="send-reminder",
    )
    def send_reminder(self, request, pk=None):
        invoice = self.get_object()

        if invoice.payment_status in [
            "PAID",
            "VOID",
        ]:
            return Response(
                {"detail": "Paid or void invoices do not require reminders."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        invoice.last_reminder_sent_at = timezone.now()

        invoice.save(
            update_fields=[
                "last_reminder_sent_at",
                "updated_at",
            ]
        )

        return Response(
            {
                "success": True,
                "message": "Payment reminder queued.",
                "data": self.get_serializer(
                    invoice,
                ).data,
            }
        )

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
    )
    def void(self, request, pk=None):
        invoice = self.get_object()

        if invoice.paid_amount > 0:
            return Response(
                {
                    "detail": "Invoices with payments cannot be voided until payments are reversed."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        invoice.payment_status = "VOID"
        invoice.balance_due = Decimal("0")
        invoice.voided_at = timezone.now()

        invoice.save(
            update_fields=[
                "payment_status",
                "balance_due",
                "voided_at",
                "updated_at",
            ]
        )

        return Response(
            self.get_serializer(
                invoice,
            ).data
        )


class POSSaleViewSet(Base):
    queryset = POSSale.objects.select_related(
        "customer",
        "branch",
        "cashier",
    ).prefetch_related(
        "items__product",
        "items__variant",
    )

    serializer_class = POSSaleSerializer

    search_fields = [
        "receipt_number",
        "customer__customer_name",
        "cashier__username",
        "cashier__first_name",
        "cashier__last_name",
        "status",
    ]

    filterset_fields = [
        "branch",
        "customer",
        "cashier",
        "payment_method",
        "status",
    ]

    ordering_fields = [
        "receipt_number",
        "sale_datetime",
        "total_amount",
        "status",
        "created_at",
    ]

    ordering = [
        "-sale_datetime",
        "-id",
    ]

    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        customers = Customer.objects.filter(
            is_active=True,
        ).order_by("customer_name")

        cashiers = User.objects.filter(
            is_active=True,
        ).order_by("first_name", "username")

        product_filters = {
            "is_active": True,
        }

        product_field_names = {field.name for field in Product._meta.get_fields()}

        if "is_deleted" in product_field_names:
            product_filters["is_deleted"] = False

        products = Product.objects.filter(
            **product_filters,
        ).order_by("product_name")

        stocks = ProductStock.objects.select_related(
            "product",
            "variant",
            "branch",
        )

        if branch_id:
            stocks = stocks.filter(
                branch_id=branch_id,
            )

        return Response(
            {
                "customers": [
                    {
                        "id": customer.id,
                        "customer_name": customer.customer_name,
                    }
                    for customer in customers
                ],
                "cashiers": [
                    {
                        "id": user.id,
                        "display_name": (user.get_full_name() or user.username),
                    }
                    for user in cashiers
                ],
                "products": [
                    {
                        "id": product.id,
                        "product_name": product.product_name,
                        "sku": getattr(
                            product,
                            "sku",
                            "",
                        ),
                        "barcode": getattr(
                            product,
                            "barcode",
                            "",
                        ),
                        "description": getattr(
                            product,
                            "description",
                            "",
                        ),
                        "selling_price": getattr(
                            product,
                            "selling_price",
                            0,
                        ),
                        "is_active": getattr(
                            product,
                            "is_active",
                            True,
                        ),
                        "is_deleted": getattr(
                            product,
                            "is_deleted",
                            False,
                        ),
                    }
                    for product in products
                ],
                "stock": [
                    {
                        "product_id": stock.product_id,
                        "variant_id": stock.variant_id,
                        "branch_id": stock.branch_id,
                        "available_stock": (stock.current_stock - stock.reserved_stock),
                    }
                    for stock in stocks
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        today = timezone.localdate()

        today_queryset = queryset.filter(
            sale_datetime__date=today,
            status="PAID",
        )

        totals = today_queryset.aggregate(
            total=Sum("total_amount"),
            average=Avg("total_amount"),
        )

        cash_total = today_queryset.filter(
            payment_method="CASH",
        ).aggregate(
            value=Sum("total_amount")
        )["value"] or Decimal("0")

        card_total = today_queryset.filter(
            payment_method="CARD",
        ).aggregate(
            value=Sum("total_amount")
        )["value"] or Decimal("0")

        split_cash = today_queryset.filter(
            payment_method="SPLIT",
        ).aggregate(
            value=Sum("cash_amount")
        )["value"] or Decimal("0")

        split_card = today_queryset.filter(
            payment_method="SPLIT",
        ).aggregate(
            value=Sum("card_amount")
        )["value"] or Decimal("0")

        cash_value = cash_total + split_cash
        card_value = card_total + split_card
        payment_total = cash_value + card_value

        return Response(
            {
                "todays_sales": totals["total"] or 0,
                "transactions": today_queryset.count(),
                "avg_basket_size": totals["average"] or 0,
                "cash_percentage": (
                    round(
                        cash_value / payment_total * 100,
                        1,
                    )
                    if payment_total
                    else 0
                ),
                "card_percentage": (
                    round(
                        card_value / payment_total * 100,
                        1,
                    )
                    if payment_total
                    else 0
                ),
                "returns_today": SalesReturn.objects.filter(
                    return_date=today,
                ).count(),
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = 'attachment; filename="pos-sales.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Receipt Number",
                "Date/Time",
                "Branch",
                "Cashier",
                "Customer",
                "Items",
                "Payment Method",
                "Subtotal",
                "VAT",
                "Discount",
                "Total",
                "Status",
            ]
        )

        for sale in queryset:
            writer.writerow(
                [
                    sale.receipt_number,
                    sale.sale_datetime or "",
                    (sale.branch.branch_name if sale.branch else ""),
                    (
                        sale.cashier.get_full_name() or sale.cashier.username
                        if sale.cashier
                        else ""
                    ),
                    (
                        sale.customer.customer_name
                        if sale.customer
                        else "Walk-in Customer"
                    ),
                    sale.items.count(),
                    sale.get_payment_method_display(),
                    sale.subtotal or 0,
                    sale.vat_amount or 0,
                    sale.discount_amount or 0,
                    sale.total_amount or 0,
                    sale.status,
                ]
            )

        return response

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
    )
    def void(self, request, pk=None):
        sale = self.get_object()

        if sale.status == "VOID":
            return Response(self.get_serializer(sale).data)

        for item in sale.items.select_related(
            "product",
            "variant",
        ):
            stock = ProductStock.objects.select_for_update().get(
                product=item.product,
                variant=item.variant,
                branch=sale.branch,
            )

            previous_stock = stock.current_stock
            stock.current_stock = stock.current_stock + item.quantity

            stock.save(
                update_fields=[
                    "current_stock",
                    "updated_at",
                ]
            )

            StockMovement.objects.create(
                movement_number=(f"MOV-VOID-{sale.receipt_number}-{item.id}"),
                product=item.product,
                variant=item.variant,
                branch=sale.branch,
                movement_type="SALE_RETURN",
                quantity=item.quantity,
                previous_stock=previous_stock,
                new_stock=stock.current_stock,
                reference_type="POS_VOID",
                reference_id=str(sale.id),
                remarks=(f"Void POS sale {sale.receipt_number}"),
                performed_by=request.user,
            )

        sale.status = "VOID"
        sale.voided_at = timezone.now()
        sale.void_reason = request.data.get(
            "reason",
            "",
        )

        sale.save(
            update_fields=[
                "status",
                "voided_at",
                "void_reason",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(sale).data)


class SalesCreditNoteViewSet(Base):
    queryset = SalesCreditNote.objects.select_related(
        "customer",
        "branch",
        "invoice",
    ).prefetch_related(
        "items__product",
        "items__variant",
        "items__invoice_item",
    )

    serializer_class = SalesCreditNoteSerializer

    search_fields = [
        "credit_note_number",
        "customer__customer_name",
        "invoice__invoice_number",
        "reason",
        "status",
    ]

    filterset_fields = [
        "branch",
        "customer",
        "invoice",
        "reason",
        "refund_method",
        "status",
    ]

    ordering_fields = [
        "credit_note_number",
        "credit_date",
        "total_amount",
        "status",
        "created_at",
        "customer__customer_name",
    ]

    ordering = [
        "-credit_date",
        "-id",
    ]

    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        invoices = (
            SalesInvoice.objects.select_related(
                "customer",
                "branch",
            )
            .exclude(
                payment_status="VOID",
            )
            .order_by(
                "-invoice_date",
                "-id",
            )
        )

        if branch_id:
            invoices = invoices.filter(
                branch_id=branch_id,
            )

        return Response(
            {
                "invoices": [
                    {
                        "id": invoice.id,
                        "invoice_number": invoice.invoice_number,
                        "customer_name": (
                            invoice.customer.customer_name if invoice.customer else ""
                        ),
                        "branch_id": invoice.branch_id,
                        "total_amount": invoice.total_amount,
                        "balance_due": invoice.balance_due,
                    }
                    for invoice in invoices
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
        url_path=r"invoice-options/(?P<invoice_id>[^/.]+)",
    )
    def invoice_options(
        self,
        request,
        invoice_id=None,
    ):
        invoice = (
            SalesInvoice.objects.select_related(
                "customer",
                "branch",
            )
            .prefetch_related(
                "items__product",
                "items__variant",
            )
            .get(pk=invoice_id)
        )

        already_credited = invoice.credit_notes.filter(status="ISSUED").aggregate(
            value=Sum("total_amount")
        )["value"] or Decimal("0")

        linked_return = invoice.returns.order_by("-id").first()

        items = []

        for item in invoice.items.all():
            credited_quantity = SalesCreditNoteItem.objects.filter(
                invoice_item=item,
                credit_note__status="ISSUED",
            ).aggregate(value=Sum("credit_quantity"))["value"] or Decimal("0")

            available_quantity = max(
                Decimal("0"),
                item.quantity - credited_quantity,
            )

            items.append(
                {
                    "id": item.id,
                    "product_id": item.product_id,
                    "variant_id": item.variant_id,
                    "product_name": (item.product.product_name if item.product else ""),
                    "description": item.description,
                    "invoiced_quantity": item.quantity,
                    "already_credited_quantity": credited_quantity,
                    "available_quantity": available_quantity,
                    "unit_price": item.unit_price,
                    "vat_percentage": item.vat_percentage,
                }
            )

        return Response(
            {
                "invoice_id": invoice.id,
                "invoice_number": invoice.invoice_number,
                "customer_id": invoice.customer_id,
                "customer_name": (
                    invoice.customer.customer_name if invoice.customer else ""
                ),
                "branch_id": invoice.branch_id,
                "invoice_total": invoice.total_amount,
                "already_credited": already_credited,
                "linked_return_number": (
                    linked_return.return_number if linked_return else ""
                ),
                "items": items,
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        today = timezone.localdate()

        issued_this_month = queryset.filter(
            status="ISSUED",
            issued_at__year=today.year,
            issued_at__month=today.month,
        )

        linked_to_returns = queryset.filter(
            invoice__returns__isnull=False,
        ).distinct()

        issued_with_duration = queryset.filter(
            status="ISSUED",
            issued_at__isnull=False,
        )

        durations = [
            (credit_note.issued_at.date() - credit_note.created_at.date()).days
            for credit_note in issued_with_duration
        ]

        return Response(
            {
                "open_credit_notes": queryset.filter(
                    status="DRAFT",
                ).count(),
                "value_issued_mtd": issued_this_month.aggregate(
                    value=Sum("total_amount")
                )["value"]
                or 0,
                "linked_to_returns": linked_to_returns.count(),
                "total_credit_notes": queryset.count(),
                "avg_processing_days": (
                    round(
                        sum(durations) / len(durations),
                        1,
                    )
                    if durations
                    else 0
                ),
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = (
            'attachment; filename="sales-credit-notes.csv"'
        )

        writer = csv.writer(response)

        writer.writerow(
            [
                "Credit Note Number",
                "Customer",
                "Invoice",
                "Credit Date",
                "Reason",
                "Refund Method",
                "Subtotal",
                "VAT",
                "Total",
                "Status",
            ]
        )

        for credit_note in queryset:
            writer.writerow(
                [
                    credit_note.credit_note_number,
                    (
                        credit_note.customer.customer_name
                        if credit_note.customer
                        else ""
                    ),
                    (credit_note.invoice.invoice_number if credit_note.invoice else ""),
                    credit_note.credit_date or "",
                    credit_note.get_reason_display(),
                    credit_note.get_refund_method_display(),
                    credit_note.subtotal or 0,
                    credit_note.vat_amount or 0,
                    credit_note.total_amount or 0,
                    credit_note.get_status_display(),
                ]
            )

        return response

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
    )
    def issue(self, request, pk=None):
        credit_note = self.get_object()

        if credit_note.status != "DRAFT":
            return Response(
                {"detail": "Only draft credit notes can be issued."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        credit_note.status = "ISSUED"
        credit_note.issued_at = timezone.now()

        credit_note.save(
            update_fields=[
                "status",
                "issued_at",
                "updated_at",
            ]
        )

        serializer = self.get_serializer(credit_note)

        serializer._apply_credit(credit_note)

        return Response(serializer.data)

    @transaction.atomic
    @action(
        detail=True,
        methods=["post"],
    )
    def void(self, request, pk=None):
        credit_note = self.get_object()

        if credit_note.status == "VOID":
            return Response(self.get_serializer(credit_note).data)

        if credit_note.status == "REFUNDED":
            return Response(
                {"detail": "Refunded credit notes cannot be voided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        credit_note.status = "VOID"
        credit_note.voided_at = timezone.now()
        credit_note.void_reason = request.data.get(
            "reason",
            "",
        )

        credit_note.save(
            update_fields=[
                "status",
                "voided_at",
                "void_reason",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(credit_note).data)


class SalesPaymentViewSet(Base):
    queryset = SalesPayment.objects.select_related(
        "customer",
        "invoice",
        "branch",
        "bank_account",
        "cash_register",
    )
    serializer_class = SalesPaymentSerializer
    search_fields = [
        "payment_number",
        "customer__customer_name",
        "invoice__invoice_number",
        "reference_number",
    ]
    filterset_fields = [
        "branch",
        "customer",
        "status",
        "payment_method",
    ]
    ordering_fields = [
        "payment_number",
        "payment_date",
        "amount",
        "status",
        "created_at",
    ]
    ordering = ["-payment_date", "-id"]

    @action(detail=False, methods=["get"], url_path="form-options")
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        invoices = (
            SalesInvoice.objects.select_related("customer", "branch")
            .filter(balance_due__gt=0)
            .exclude(payment_status="VOID")
            .order_by("-invoice_date", "-id")
        )
        bank_accounts = BankAccount.objects.filter(is_active=True)
        cash_registers = CashRegister.objects.filter(is_active=True)

        if branch_id:
            invoices = invoices.filter(branch_id=branch_id)

            bank_fields = {field.name for field in BankAccount._meta.get_fields()}
            register_fields = {field.name for field in CashRegister._meta.get_fields()}

            if "branch" in bank_fields:
                bank_accounts = bank_accounts.filter(branch_id=branch_id)
            if "branch" in register_fields:
                cash_registers = cash_registers.filter(branch_id=branch_id)

        return Response(
            {
                "invoices": [
                    {
                        "id": invoice.id,
                        "invoice_number": invoice.invoice_number,
                        "customer_name": (
                            invoice.customer.customer_name if invoice.customer else ""
                        ),
                    }
                    for invoice in invoices
                ],
                "bank_accounts": [
                    {
                        "id": account.id,
                        "account_name": getattr(
                            account,
                            "account_name",
                            str(account),
                        ),
                    }
                    for account in bank_accounts
                ],
                "cash_registers": [
                    {
                        "id": register.id,
                        "name": getattr(
                            register,
                            "name",
                            str(register),
                        ),
                    }
                    for register in cash_registers
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
        url_path=r"invoice-options/(?P<invoice_id>[^/.]+)",
    )
    def invoice_options(self, request, invoice_id=None):
        invoice = SalesInvoice.objects.select_related(
            "customer",
            "branch",
        ).get(pk=invoice_id)

        return Response(
            {
                "invoice_id": invoice.id,
                "invoice_number": invoice.invoice_number,
                "customer_id": invoice.customer_id,
                "customer_name": (
                    invoice.customer.customer_name if invoice.customer else ""
                ),
                "branch_id": invoice.branch_id,
                "currency": invoice.currency,
                "invoice_total": invoice.total_amount,
                "paid_amount": invoice.paid_amount,
                "balance_due": invoice.balance_due,
            }
        )

    @action(detail=False, methods=["get"])
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        today = timezone.localdate()

        return Response(
            {
                "received_today": (
                    queryset.filter(
                        payment_date=today,
                        status="PAID",
                    ).aggregate(
                        value=Sum("amount")
                    )["value"]
                    or 0
                ),
                "received_mtd": (
                    queryset.filter(
                        payment_date__year=today.year,
                        payment_date__month=today.month,
                        status="PAID",
                    ).aggregate(value=Sum("amount"))["value"]
                    or 0
                ),
                "pending_clearance": (
                    queryset.filter(status="PENDING").aggregate(value=Sum("amount"))[
                        "value"
                    ]
                    or 0
                ),
                "active_methods": (
                    queryset.values("payment_method").distinct().count()
                ),
            }
        )

    @action(detail=False, methods=["get"])
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sales-payments.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "Payment Number",
                "Customer",
                "Invoice",
                "Payment Date",
                "Method",
                "Amount",
                "Currency",
                "Reference",
                "Status",
            ]
        )

        for payment in queryset:
            writer.writerow(
                [
                    payment.payment_number,
                    payment.customer.customer_name if payment.customer else "",
                    payment.invoice.invoice_number if payment.invoice else "",
                    payment.payment_date or "",
                    payment.get_payment_method_display(),
                    payment.amount or 0,
                    payment.currency or "AED",
                    payment.reference_number or "",
                    payment.get_status_display(),
                ]
            )
        return response


class SalesReturnViewSet(Base):
    queryset = SalesReturn.objects.select_related(
        "customer",
        "branch",
        "sales_order",
        "invoice",
        "approved_by",
    ).prefetch_related(
        "items__product",
        "items__variant",
        "items__sales_order_item",
    )
    serializer_class = SalesReturnSerializer
    search_fields = [
        "return_number",
        "customer__customer_name",
        "sales_order__order_number",
        "invoice__invoice_number",
    ]
    filterset_fields = [
        "branch",
        "customer",
        "status",
        "reason",
        "resolution",
    ]
    ordering_fields = [
        "return_number",
        "return_date",
        "total_amount",
        "status",
        "created_at",
    ]
    ordering = ["-return_date", "-id"]

    @action(detail=False, methods=["get"], url_path="form-options")
    def form_options(self, request):
        branch_id = request.query_params.get("branch")

        orders = (
            SalesOrder.objects.select_related("customer", "branch")
            .exclude(status="CANCELLED")
            .order_by("-order_date", "-id")
        )

        if branch_id:
            orders = orders.filter(branch_id=branch_id)

        return Response(
            {
                "sales_orders": [
                    {
                        "id": order.id,
                        "order_number": order.order_number,
                        "customer_name": (
                            order.customer.customer_name if order.customer else ""
                        ),
                    }
                    for order in orders
                ]
            }
        )

    @action(
        detail=False,
        methods=["get"],
        url_path=r"order-options/(?P<order_id>[^/.]+)",
    )
    def order_options(self, request, order_id=None):
        order = (
            SalesOrder.objects.select_related("customer", "branch")
            .prefetch_related(
                "items__product",
                "items__variant",
                "invoices",
            )
            .get(pk=order_id)
        )
        invoice = order.invoices.order_by("-id").first()
        items = []

        for order_item in order.items.all():
            already_returned = SalesReturnItem.objects.filter(
                sales_order_item=order_item,
            ).exclude(
                sales_return__status__in=[
                    "REJECTED",
                    "CANCELLED",
                ]
            ).aggregate(
                value=Sum("returned_quantity")
            )[
                "value"
            ] or Decimal(
                "0"
            )

            items.append(
                {
                    "id": order_item.id,
                    "product_id": order_item.product_id,
                    "variant_id": order_item.variant_id,
                    "product_name": (
                        order_item.product.product_name if order_item.product else ""
                    ),
                    "description": order_item.description,
                    "ordered_quantity": order_item.quantity,
                    "already_returned_quantity": already_returned,
                    "available_quantity": max(
                        Decimal("0"),
                        order_item.quantity - already_returned,
                    ),
                    "unit_price": order_item.unit_price,
                }
            )

        return Response(
            {
                "order_id": order.id,
                "order_number": order.order_number,
                "customer_id": order.customer_id,
                "customer_name": (
                    order.customer.customer_name if order.customer else ""
                ),
                "branch_id": order.branch_id,
                "order_total": order.total_amount,
                "invoice_id": invoice.id if invoice else None,
                "invoice_number": invoice.invoice_number if invoice else "",
                "already_returned_value": (
                    order.returns.exclude(
                        status__in=["REJECTED", "CANCELLED"]
                    ).aggregate(value=Sum("total_amount"))["value"]
                    or 0
                ),
                "items": items,
            }
        )

    @action(detail=False, methods=["get"])
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        today = timezone.localdate()
        completed = queryset.filter(
            status="COMPLETED",
            completed_at__isnull=False,
        )
        resolution_days = [
            (item.completed_at.date() - item.created_at.date()).days
            for item in completed
        ]

        return Response(
            {
                "open_returns": queryset.filter(
                    status__in=[
                        "DRAFT",
                        "PENDING_APPROVAL",
                        "APPROVED",
                    ]
                ).count(),
                "value_mtd": (
                    queryset.filter(
                        return_date__year=today.year,
                        return_date__month=today.month,
                    ).aggregate(value=Sum("total_amount"))["value"]
                    or 0
                ),
                "restocked": queryset.filter(
                    status="COMPLETED",
                    items__condition="SELLABLE",
                )
                .distinct()
                .count(),
                "avg_resolution_days": (
                    round(
                        sum(resolution_days) / len(resolution_days),
                        1,
                    )
                    if resolution_days
                    else 0
                ),
            }
        )

    @action(detail=False, methods=["get"])
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sales-returns.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "Return Number",
                "Customer",
                "Sales Order",
                "Invoice",
                "Date",
                "Reason",
                "Resolution",
                "Total",
                "Status",
            ]
        )
        for item in queryset:
            writer.writerow(
                [
                    item.return_number,
                    item.customer.customer_name if item.customer else "",
                    item.sales_order.order_number if item.sales_order else "",
                    item.invoice.invoice_number if item.invoice else "",
                    item.return_date or "",
                    item.get_reason_display(),
                    item.get_resolution_display(),
                    item.total_amount or 0,
                    item.get_status_display(),
                ]
            )
        return response


class PriceListViewSet(Base):
    queryset = PriceList.objects.select_related("branch").prefetch_related(
        "items", "customers"
    )
    serializer_class = PriceListSerializer
    search_fields = [
        "name",
        "status",
        "customer_category",
    ]
    filterset_fields = [
        "branch",
        "status",
        "applies_to",
        "discount_type",
    ]
    ordering_fields = [
        "name",
        "valid_from",
        "valid_until",
        "status",
        "created_at",
    ]
    ordering = ["-created_at", "-id"]

    @action(detail=False, methods=["get"], url_path="form-options")
    def form_options(self, request):
        customers = Customer.objects.filter(
            is_active=True,
        ).order_by("customer_name")

        product_filters = {"is_active": True}
        product_fields = {field.name for field in Product._meta.get_fields()}
        if "is_deleted" in product_fields:
            product_filters["is_deleted"] = False

        products = Product.objects.filter(**product_filters).order_by("product_name")

        return Response(
            {
                "customers": [
                    {
                        "id": customer.id,
                        "customer_name": customer.customer_name,
                        "category": getattr(customer, "category", ""),
                    }
                    for customer in customers
                ],
                "products": [
                    {
                        "id": product.id,
                        "product_name": product.product_name,
                        "sku": getattr(product, "sku", ""),
                        "selling_price": getattr(
                            product,
                            "selling_price",
                            0,
                        ),
                    }
                    for product in products
                ],
            }
        )

    @action(detail=False, methods=["get"])
    def summary(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        today = timezone.localdate()
        week_end = today + timedelta(days=7)

        active_promotions = queryset.filter(
            status="ACTIVE",
            valid_from__lte=today,
        ).filter(Q(valid_until__isnull=True) | Q(valid_until__gte=today))

        return Response(
            {
                "active_price_lists": queryset.filter(status="ACTIVE").count(),
                "items_with_overrides": PriceListItem.objects.filter(
                    price_list__in=queryset,
                    custom_price__isnull=False,
                ).count(),
                "active_promotions": active_promotions.count(),
                "expiring_this_week": queryset.filter(
                    status="ACTIVE",
                    valid_until__range=[today, week_end],
                ).count(),
            }
        )

    @action(detail=False, methods=["get"])
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="price-lists.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "Name",
                "Applies To",
                "Discount Type",
                "Discount",
                "Valid From",
                "Valid Until",
                "Status",
            ]
        )
        for item in queryset:
            discount = (
                f"{item.discount_percentage}%"
                if item.discount_type == "PERCENTAGE"
                else (
                    f"AED {item.fixed_discount}"
                    if item.discount_type == "FIXED"
                    else "Custom prices"
                )
            )
            writer.writerow(
                [
                    item.name,
                    item.get_applies_to_display(),
                    item.get_discount_type_display(),
                    discount,
                    item.valid_from or "",
                    item.valid_until or "",
                    item.get_status_display(),
                ]
            )
        return response


class SalesReportViewSet(Base):
    queryset = SalesReport.objects.select_related(
        "branch",
        "customer",
    )

    serializer_class = SalesReportSerializer

    search_fields = [
        "report_name",
        "report_type",
        "owner_team",
        "email_to",
        "status",
    ]

    filterset_fields = [
        "branch",
        "report_type",
        "period",
        "group_by",
        "sales_channel",
        "output_format",
        "recurrence",
        "status",
    ]

    ordering_fields = [
        "report_name",
        "report_type",
        "period",
        "generated_at",
        "output_format",
        "owner_team",
        "status",
        "created_at",
    ]

    ordering = [
        "-created_at",
        "-id",
    ]

    def _date_range(self, report):
        today = timezone.localdate()

        if report.period == "THIS_MONTH":
            start = today.replace(day=1)
            end = today

        elif report.period == "THIS_QUARTER":
            quarter_start_month = (today.month - 1) // 3 * 3 + 1

            start = today.replace(
                month=quarter_start_month,
                day=1,
            )

            end = today

        elif report.period == "THIS_YEAR":
            start = today.replace(
                month=1,
                day=1,
            )

            end = today

        else:
            start = report.custom_start
            end = report.custom_end

        return start, end

    def _invoice_queryset(self, report):
        start, end = self._date_range(report)

        queryset = (
            SalesInvoice.objects.select_related(
                "customer",
                "branch",
                "salesperson",
            )
            .prefetch_related(
                "items__product",
            )
            .exclude(
                payment_status="VOID",
            )
        )

        if start:
            queryset = queryset.filter(
                invoice_date__gte=start,
            )

        if end:
            queryset = queryset.filter(
                invoice_date__lte=end,
            )

        if report.branch_id:
            queryset = queryset.filter(
                branch_id=report.branch_id,
            )

        if report.customer_id:
            queryset = queryset.filter(
                customer_id=report.customer_id,
            )

        if report.sales_channel == "ORDER":
            queryset = queryset.filter(
                sales_order__isnull=False,
            )

        elif report.sales_channel == "INVOICE":
            queryset = queryset.filter(
                sales_order__isnull=True,
            )

        return queryset

    def _report_rows(self, report):
        invoices = self._invoice_queryset(report)

        if report.group_by == "CUSTOMER":
            data = (
                invoices.values(
                    "customer_id",
                    "customer__customer_name",
                )
                .annotate(
                    invoice_count=Count("id"),
                    sales_total=Sum("total_amount"),
                    paid_total=Sum("paid_amount"),
                    balance_total=Sum("balance_due"),
                )
                .order_by("-sales_total")
            )

            return [
                {
                    "Customer": row["customer__customer_name"]
                    or "Walk-in / Unassigned",
                    "Invoices": row["invoice_count"],
                    "Sales": row["sales_total"] or 0,
                    "Paid": row["paid_total"] or 0,
                    "Balance": row["balance_total"] or 0,
                }
                for row in data
            ]

        if report.group_by == "PRODUCT":
            data = (
                SalesInvoiceItem.objects.filter(
                    invoice__in=invoices,
                )
                .values(
                    "product_id",
                    "product__product_name",
                )
                .annotate(
                    quantity=Sum("quantity"),
                    sales_total=Sum("line_total"),
                )
                .order_by("-sales_total")
            )

            return [
                {
                    "Product": row["product__product_name"] or "Unassigned Product",
                    "Quantity": row["quantity"] or 0,
                    "Sales": row["sales_total"] or 0,
                }
                for row in data
            ]

        if report.group_by == "CASHIER":
            data = POSSale.objects.filter(
                status="PAID",
            )

            start, end = self._date_range(report)

            if start:
                data = data.filter(
                    sale_datetime__date__gte=start,
                )

            if end:
                data = data.filter(
                    sale_datetime__date__lte=end,
                )

            if report.branch_id:
                data = data.filter(
                    branch_id=report.branch_id,
                )

            data = (
                data.values(
                    "cashier_id",
                    "cashier__first_name",
                    "cashier__last_name",
                    "cashier__username",
                )
                .annotate(
                    transactions=Count("id"),
                    sales_total=Sum("total_amount"),
                )
                .order_by("-sales_total")
            )

            return [
                {
                    "Cashier": (
                        " ".join(
                            filter(
                                None,
                                [
                                    row["cashier__first_name"],
                                    row["cashier__last_name"],
                                ],
                            )
                        )
                        or row["cashier__username"]
                        or "Unassigned"
                    ),
                    "Transactions": row["transactions"],
                    "Sales": row["sales_total"] or 0,
                }
                for row in data
            ]

        if report.group_by == "PAYMENT_METHOD":
            start, end = self._date_range(report)

            payments = SalesPayment.objects.filter(
                status="PAID",
            )

            if start:
                payments = payments.filter(
                    payment_date__gte=start,
                )

            if end:
                payments = payments.filter(
                    payment_date__lte=end,
                )

            if report.branch_id:
                payments = payments.filter(
                    branch_id=report.branch_id,
                )

            data = (
                payments.values(
                    "payment_method",
                )
                .annotate(
                    transactions=Count("id"),
                    amount=Sum("amount"),
                )
                .order_by("-amount")
            )

            labels = dict(SalesPayment.PAYMENT_METHOD_CHOICES)

            return [
                {
                    "Payment Method": labels.get(
                        row["payment_method"],
                        row["payment_method"],
                    ),
                    "Transactions": row["transactions"],
                    "Amount": row["amount"] or 0,
                }
                for row in data
            ]

        data = (
            invoices.values(
                "sale_type",
            )
            .annotate(
                invoice_count=Count("id"),
                sales_total=Sum("total_amount"),
            )
            .order_by("-sales_total")
        )

        labels = dict(SalesInvoice.SALE_TYPE_CHOICES)

        return [
            {
                "Sales Channel": labels.get(
                    row["sale_type"],
                    row["sale_type"],
                ),
                "Invoices": row["invoice_count"],
                "Sales": row["sales_total"] or 0,
            }
            for row in data
        ]

    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="form-options",
    )
    def form_options(self, request):
        customer_filters = {}

        customer_fields = {field.name for field in Customer._meta.get_fields()}

        if "is_active" in customer_fields:
            customer_filters["is_active"] = True

        if "is_deleted" in customer_fields:
            customer_filters["is_deleted"] = False

        customers = Customer.objects.filter(**customer_filters).order_by(
            "customer_name"
        )

        return Response(
            {
                "customers": [
                    {
                        "id": customer.id,
                        "customer_name": customer.customer_name,
                    }
                    for customer in customers
                ],
                "owner_teams": [
                    {
                        "value": "Finance Team",
                        "label": "Finance Team",
                    },
                    {
                        "value": "Sales Team",
                        "label": "Sales Team",
                    },
                    {
                        "value": "Management",
                        "label": "Management",
                    },
                    {
                        "value": "Branch Managers",
                        "label": "Branch Managers",
                    },
                ],
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def summary(self, request):
        branch_id = request.query_params.get("branch")

        today = timezone.localdate()
        month_start = today.replace(
            day=1,
        )

        invoices = SalesInvoice.objects.exclude(
            payment_status="VOID",
        ).filter(
            invoice_date__gte=month_start,
            invoice_date__lte=today,
        )

        orders = SalesOrder.objects.filter(
            order_date__gte=month_start,
            order_date__lte=today,
        )

        quotations = Quotation.objects.filter(
            quote_date__gte=month_start,
            quote_date__lte=today,
        )

        if branch_id:
            invoices = invoices.filter(
                branch_id=branch_id,
            )

            orders = orders.filter(
                branch_id=branch_id,
            )

            quotations = quotations.filter(
                branch_id=branch_id,
            )

        top_customer = (
            invoices.values(
                "customer__customer_name",
            )
            .annotate(
                value=Sum("total_amount"),
            )
            .order_by("-value")
            .first()
        )

        order_count = orders.count()
        quotation_count = quotations.count()

        return Response(
            {
                "revenue_mtd": invoices.aggregate(value=Sum("total_amount"))["value"]
                or 0,
                "revenue_change": 0,
                "orders_mtd": order_count,
                "orders_change": 0,
                "top_customer": (
                    top_customer["customer__customer_name"] if top_customer else ""
                ),
                "top_customer_value": (top_customer["value"] if top_customer else 0),
                "conversion_rate": (
                    round(
                        order_count / quotation_count * 100,
                        1,
                    )
                    if quotation_count
                    else 0
                ),
            }
        )

    @action(
        detail=False,
        methods=["get"],
    )
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        response = HttpResponse(
            content_type="text/csv",
        )

        response["Content-Disposition"] = 'attachment; filename="sales-reports.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Report",
                "Type",
                "Period",
                "Generated",
                "Format",
                "Owner",
                "Recurrence",
                "Status",
            ]
        )

        for report in queryset:
            writer.writerow(
                [
                    report.report_name,
                    report.get_report_type_display(),
                    report.get_period_display(),
                    report.generated_at or "",
                    report.get_output_format_display(),
                    report.owner_team or "",
                    report.get_recurrence_display(),
                    report.get_status_display(),
                ]
            )

        return response

    @action(
        detail=True,
        methods=["get"],
    )
    def download(self, request, pk=None):
        report = self.get_object()
        rows = self._report_rows(report)

        if not rows:
            rows = [{"Message": "No data found for the selected report filters."}]

        if report.output_format == "CSV":
            response = HttpResponse(
                content_type="text/csv",
            )

            response["Content-Disposition"] = (
                f'attachment; filename="sales-report-{report.id}.csv"'
            )

            writer = csv.DictWriter(
                response,
                fieldnames=list(rows[0].keys()),
            )

            writer.writeheader()
            writer.writerows(rows)

            return response

        if report.output_format == "EXCEL":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Sales Report"

                headers = list(rows[0].keys())

                sheet.append(headers)

                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                for row in rows:
                    sheet.append(
                        [
                            row.get(
                                header,
                                "",
                            )
                            for header in headers
                        ]
                    )

                for index, header in enumerate(
                    headers,
                    start=1,
                ):
                    sheet.column_dimensions[get_column_letter(index)].width = max(
                        15,
                        len(str(header)) + 3,
                    )

                buffer = BytesIO()
                workbook.save(buffer)
                buffer.seek(0)

                response = HttpResponse(
                    buffer.getvalue(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )

                response["Content-Disposition"] = (
                    f'attachment; filename="sales-report-{report.id}.xlsx"'
                )

                return response

            except ImportError:
                return Response(
                    {"detail": "Install openpyxl to generate Excel reports."},
                    status=status.HTTP_501_NOT_IMPLEMENTED,
                )

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )

            buffer = BytesIO()

            document = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                rightMargin=24,
                leftMargin=24,
                topMargin=24,
                bottomMargin=24,
            )

            styles = getSampleStyleSheet()

            elements = [
                Paragraph(
                    report.report_name,
                    styles["Title"],
                ),
                Paragraph(
                    (
                        f"{report.get_period_display()} · "
                        f"Grouped by {report.get_group_by_display()}"
                    ),
                    styles["Normal"],
                ),
                Spacer(
                    1,
                    14,
                ),
            ]

            headers = list(rows[0].keys())

            table_data = [headers] + [
                [
                    str(
                        row.get(
                            header,
                            "",
                        )
                    )
                    for header in headers
                ]
                for row in rows
            ]

            table = Table(
                table_data,
                repeatRows=1,
            )

            table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (
                                0,
                                0,
                            ),
                            (
                                -1,
                                0,
                            ),
                            colors.HexColor("#2563EB"),
                        ),
                        (
                            "TEXTCOLOR",
                            (
                                0,
                                0,
                            ),
                            (
                                -1,
                                0,
                            ),
                            colors.white,
                        ),
                        (
                            "FONTNAME",
                            (
                                0,
                                0,
                            ),
                            (
                                -1,
                                0,
                            ),
                            "Helvetica-Bold",
                        ),
                        (
                            "GRID",
                            (
                                0,
                                0,
                            ),
                            (
                                -1,
                                -1,
                            ),
                            0.5,
                            colors.grey,
                        ),
                        (
                            "VALIGN",
                            (
                                0,
                                0,
                            ),
                            (
                                -1,
                                -1,
                            ),
                            "MIDDLE",
                        ),
                        (
                            "ROWBACKGROUNDS",
                            (
                                0,
                                1,
                            ),
                            (
                                -1,
                                -1,
                            ),
                            [
                                colors.white,
                                colors.HexColor("#F8FAFC"),
                            ],
                        ),
                    ]
                )
            )

            elements.append(table)

            document.build(elements)

            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/pdf",
            )

            response["Content-Disposition"] = (
                f'attachment; filename="sales-report-{report.id}.pdf"'
            )

            return response

        except ImportError:
            return Response(
                {"detail": "Install reportlab to generate PDF reports."},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

    @action(
        detail=True,
        methods=["post"],
        url_path="run-now",
    )
    def run_now(self, request, pk=None):
        report = self.get_object()

        report.status = "READY"
        report.generated_at = timezone.now()
        report.last_run_at = timezone.now()
        report.error_message = None

        report.save(
            update_fields=[
                "status",
                "generated_at",
                "last_run_at",
                "error_message",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(report).data)
